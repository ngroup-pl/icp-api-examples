import os
import sys
from os.path import join, dirname
import requests
from dotenv import load_dotenv
from dateutil import parser
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font

import smtplib
from os.path import basename
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formatdate

import logging

log = logging.getLogger(__name__)

load_dotenv(join(dirname(__file__), '.env'))

icp_authorization_token = os.environ.get("ICP_AUTHORIZATION_TOKEN")
icp_instance_slug = os.environ.get("ICP_INSTANCE_SLUG")

headers = {
    'X-Auth-Token': icp_authorization_token,
    'Accept': 'application/json',
    'Content-type': 'application/json'
}

icp_api_url = f"https://app.icproject.com/api/instance/{icp_instance_slug}"


def money(v):
    return str("%.2f" % v).replace(".", ",")


def retrieve_unpaid_invoices() -> list:
    page = 1
    per_page = 100
    date = datetime.today() - timedelta(days=1)
    items = []
    while True:
        log.debug(f"Retrieving unpaid invoices, page: {page}",)
        response = requests.get(
            f"{icp_api_url}/finance/invoices",
            params={
                "page": page,
                "itemsPerPage": per_page,
                "isNotPaid": 1,
                "dateDeadline[before]": date.isoformat(),
                "order[dateDeadline]": "asc"
            },
            headers=headers,
            timeout=10
        )

        items += response.json()

        if len(response.json()) < per_page:
            break

        page = page + 1

    return items


def create_xls(invoices: list) -> (str, int):
    log.debug("creating excel")
    fn = f"invoices-to-pay-{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    wb = Workbook()

    wb.remove(wb.active)
    ws_0 = wb.create_sheet("Podsumowanie", 0)
    ws_1 = wb.create_sheet("Do importu", 1)

    ws_0.append([
        "Dane do importu znajdują się na drugim arkuszu!"
    ])
    ws_0.append([])
    ws_0.merge_cells("A1:L2")

    ws_0['A1'].fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    ws_0['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_0['A1'].font = Font(color='FFFFFF', size=12)

    ws_0.append([
        'NIP',
        'Nazwa kontrahenta',
        'Kraj kontrahenta',
        'Numer faktury',
        'Kwota',
        'Kwota [pozostało]',
        'Waluta',
        'Data wymagalności',
        'Wystawił',
        'Data wystawienia',
        'Rodzaj',
        'Link'
    ])
    ws_1.append([
        'NIP',
        'Nazwa kontrahenta',
        'Kraj kontrahenta',
        'Numer faktury',
        'Kwota',
        'Kwota [pozostało]',
        'Waluta',
        'Data wymagalności',
    ])

    to_pay = 0
    for invoice in invoices:
        deadline = parser.parse(invoice['dateDeadline'])
        to_pay += invoice['toPay'] - invoice['alreadyPaid']

        ws_0.append([
            invoice['buyerVatId'],
            invoice['buyerName'],
            'PL',
            invoice['no'],
            money(invoice['toPay']),
            money(invoice['toPay'] - invoice['alreadyPaid']),
            invoice['currencyCode'],
            deadline.strftime('%d.%m.%Y'),
            invoice['sellerCreatorName'],
            parser.parse(invoice['dateIssue']).strftime('%d.%m.%Y'),
            invoice['kind'],
            f"https://app.icproject.com/{icp_instance_slug}/finance/invoice/update/{invoice['id']}"
        ])
        ws_1.append([
            invoice['buyerVatId'],
            invoice['buyerName'],
            'PL',
            invoice['no'],
            money(invoice['toPay']),
            money(invoice['toPay'] - invoice['alreadyPaid']),
            invoice['currencyCode'],
            deadline.strftime('%d.%m.%Y'),
        ])

    ws_0.append([
        "",
        "",
        "",
        "",
        "",
        money(to_pay),
        "",
        "",
        "",
        "",
    ])

    wb.save(fn)

    log.debug("excel created")

    return fn, to_pay


def send_mail(send_from, send_to, subject, text, files=None):
    log.debug("sending mail")
    msg = MIMEMultipart()
    msg['From'] = send_from
    msg['To'] = ", ".join(send_to)
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject

    msg.attach(MIMEText(text))

    for f in files or []:
        with open(f, "rb") as fil:
            part = MIMEApplication(
                fil.read(),
                Name=basename(f)
            )
        # After the file is closed
        part['Content-Disposition'] = 'attachment; filename="%s"' % basename(f)
        msg.attach(part)

    smtp = smtplib.SMTP(os.environ.get("SMTP_HOST"), int(os.environ.get("SMTP_PORT")))
    if os.environ.get("SMTP_USE_TLS"):
        smtp.starttls()
    smtp.login(os.environ.get("SMTP_USERNAME"), os.environ.get("SMTP_PASSWORD"))
    smtp.sendmail(send_from, send_to, msg.as_string())
    smtp.close()
    log.debug("mail sent")


if __name__ == '__main__':
    logging.basicConfig(filename='icp-cashcollector.log', level=logging.DEBUG, format='%(asctime)s %(message)s')
    log.debug("start")

    try:
        unpaid_invoices = retrieve_unpaid_invoices()
    except Exception as e:
        log.error(e)
        sys.exit(1)

    _len = len(unpaid_invoices)
    if _len == 0:
        log.debug("no unpaid invoices")
        sys.exit(1)

    try:
        xls_fn, to_pay = create_xls(unpaid_invoices)
    except Exception as e:
        log.error(e)
        sys.exit(1)

    content = f"Liczba nieopłaconych faktur: {_len}, na kwotę {to_pay}"

    try:
        send_mail(
            os.environ.get("SEND_FROM"),
            [os.environ.get("SEND_TO")],
            f"[ICP] Nieopłacone faktury na dzień {datetime.now().strftime('%d.%m.%Y')}",
            content,
            [xls_fn]
        )
        os.remove(xls_fn)
    except Exception as e:
        log.error(e)
        sys.exit(1)
